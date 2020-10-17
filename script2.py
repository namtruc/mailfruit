import pandas as pd
import os
import time
import sys

from tkinter import Tk    
from tkinter.filedialog import askopenfilename
from configparser import ConfigParser  
from os import path as os_path
from openpyxl import load_workbook

dossier_python = os_path.abspath(os_path.split(__file__)[0])
dossier_usr = dossier_python + '/fichiers_utilisateur'

os.chdir(dossier_python)

parser = ConfigParser() 
parser.read('configuration.ini')




def convert_char(old): ### fct convertion lettre en chiffre
	if len(old) != 1:
		return 0
	new = ord(old)
	if 65 <= new <= 90: # Majuscules
		return new - 64
	elif 97 <= new <= 122: # Minuscules   
		return new - 96 
	return 0 # Autres
	
def fct1 (lettre): ### fct verification lettre
	while convert_char(lettre) == 0 :
		print ("erreur")
		lettre = input("Indiquer de nouveau la lettre\n")
	else :
		lettre = convert_char(lettre)
		return lettre-1
		
###########################################################################3
### Donneur


input("\nAppuyer sur Entree pour choisir le fichier excel comportant les liens catalogue\n")

Tk().withdraw() 
contenu1 = askopenfilename()

xl = pd.ExcelFile(contenu1)

v_sheet1 = 0

if len(xl.sheet_names) > 1 :
	print ("\n---Les differents feuilles presentes---")
	for x in range (len(xl.sheet_names)) :
		print (str(x)+" "+xl.sheet_names[x])
	
	v_sheet1 = int(input("Entrer le numero de la feuille excel \n"))


print ("\n----Verifier que les informations ci dessous sont correctes----")
time.sleep(1)

#os.chdir(dossier_python)

vgrp1 = parser.get('donneur', 'grp')###
vlien1 = parser.get('donneur', 'lien')####

print ("Lettre colonne comportant les liens "+vlien1)
print ("Lettre colonne indiquant le groupe "+vgrp1)

while 1:
	
	reponse = input("\nTaper 1 pour modifier manuellement les colonnes prises en compte, sinon taper 2\n")

	if reponse=='1':
	
		vlien1 = input("Indiquer la lettre de la colonne comportant les liens\n")
		vlien1 = fct1 (vlien1)
		vgrp1 = input("Indiquer la lettre de la colonne comportant le groupe\n")
		vgrp1 = fct1 (vgrp1)
		break
		
	elif reponse == '2' :

		vlien1 = fct1 (vlien1)
		vgrp1 = fct1 (vgrp1)
		break
	
	else :

		print ("Choix incorrect !")


########################################
### Receveur

input("\nAppuyer sur Entree pour choisir le fichier excel comportant les groupes et mails des clients\n")

Tk().withdraw() 
contenu2 = askopenfilename()

xl = pd.ExcelFile(contenu2)

if len(xl.sheet_names) > 1 :
	print ("\n---Les differents feuilles presentes---")
	for x in range (len(xl.sheet_names)) :
		print (str(x)+" "+xl.sheet_names[x])

	y = int(input("Entrer le numero de la feuille excel\n"))
	v_sheet2 = xl.sheet_names[y]



print ("\n----Verifier que les informations ci dessous sont correctes----")
time.sleep(1)


vgrp2 = parser.get('receveur', 'grp')
vlien2 = parser.get('receveur', 'lien')

print ("Lettre colonne ou les liens seront inscris "+vlien2)
print ("Lettre colonne indiquant le groupe "+vgrp2)

while 1:
	
	reponse = input("\nTaper 1 pour modifier manuellement les colonnes prises en compte, sinon taper 2\n")

	if reponse=='1':
	
		vlien2 = input("Indiquer la lettre de la colonne comportant les liens\n")
		vlien2 = fct1 (vlien2)
		vgrp2 = input("Indiquer la lettre de la colonne comportant le groupe\n")
		vgrp2 = fct1 (vgrp2)
		break
		
	elif reponse == '2' :

		vlien2 = fct1 (vlien2)
		vgrp2 = fct1 (vgrp2)
		break
	
	else :

		print ("Choix incorrect !")



###############################################################
#### Lecture fichier donneur

i = 0


read_file = pd.read_excel(contenu1,sheet_name=v_sheet1, header=None)         

read_file.to_csv ("Test1.csv",  
                  index = None, 
                  header = True)


df1 = pd.DataFrame(pd.read_csv("Test1.csv")) 

vgrp1a = df1.columns.values[vgrp1] 

pd.DataFrame(df1.dropna(subset=[vgrp1a], inplace=True))######suppression des grp vide
df1.drop(0, inplace=True)


list1 = [] ### reconstruction de l'index

for x in range (len(df1.index)):
 list1.append(x)

df1.index = list1


vlien1a = df1.columns.values[vlien1] 
check = pd.isna(df1[vlien1a])
for x in range ((len(df1.index))):
		if check[x] == True :
			print ("Erreur case vide colonne lien groupe "+str(df1.iat[x,vgrp1]) )
			i = i + 1
			
if i == 1 :
	print (str(i)+"erreurs")
	reponse = input("Appuyer sur entree pour quitter la programme")
	sys.exit(1)

dico = dict()
dico_verif = dict()

for x in range (len(df1.index)):
	dico.update( {df1.iat[x,vgrp1] : df1.iat[x,vlien1]} )

for x in range (len(df1.index)):
	dico_verif.update( {df1.iat[x,vgrp1] : ''} )



###############################################################
#### Ecriture fichier receveur


wb2 = load_workbook(filename = contenu2)
if len(xl.sheet_names) > 1 :
	ws2 = wb2[v_sheet2]
else :
	ws2 = wb2.active

list_erreur = []

for x in range (1,ws2.max_row+1) :
	c = (ws2.cell(row=x, column=vgrp2+1)).value
		
	for key in dico : 
		if str(key).lower().strip() == str(c).lower().strip() : 
			d = dico.get(key)
			ws2.cell(row = x, column = vlien2+1).value = d
			if key in dico_verif :
				del dico_verif[key]
			
			

wb2.save (dossier_usr + '/fichie_complet.xlsx')

if dico_verif :
	print (str(len(dico_verif))+" groupe(s) non trouve(s)")
	for key in dico_verif :
		print (key)
	reponse = input("\nTaper 1 pour quitter le programme, 2 pour continuer\n")
	if reponse == '1':
		os.remove(dossier_usr + '/fichie_complet.xlsx')
		os.remove("Test1.csv")
		sys.exit(1)
	


	
########################
### Fin

print ("Fichier de sortie enregistre sous le nom --fichier_complet.xlsx-- dans le repertoire fichier_utilisateur")
time.sleep(2)

os.remove("Test1.csv")


#### Si erreur quitter l'autre script ( code exit?)
