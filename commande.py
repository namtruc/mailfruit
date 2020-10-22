import sys
import re
import os
import datetime
import pandas as pd
import time
import getpass

from tkinter import Tk    
from tkinter.filedialog import askopenfilename
from datetime import date
from os import path as os_path
from configparser import ConfigParser  
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def convert_char(old): ### fct convertion lettre en chiffre
	if len(old) != 1:
		return 0
	new = ord(old)
	if 65 <= new <= 90: # Majuscules
		return new - 64
	elif 97 <= new <= 122: # Minuscules   
		return new - 96 
	return 0 # Autres
       
def fct2 (lettre): ### fct verification lettre
	while convert_char(lettre) == 0 :
		print ("erreur")
		lettre = input("Indiquer de nouveau la lettre\n")
	else :
		lettre = convert_char(lettre)
		return lettre-1
		
def fct3 (vide, mot, variable): ###fct verif case vide
	vide = df.columns.values[vide]
	check = pd.isna(df[vide])
	for x in range (variable, (len(df.index))):
		if check[x+1] == True :
			print ("Erreur case vide colonne ---"+mot+"--- groupe "+str(df.iat[x,vgrp])+" departement "+str(df.iat[x,vdept]) )
			return (1, x+2)
	return (0,0)

					
parser = ConfigParser() 
parser.read('configuration.ini')

dossier_python = os_path.abspath(os_path.split(__file__)[0])
dossier_usr = dossier_python + '/fichiers_utilisateur'

dv = DataValidation(type="whole")

#######################################################################################################
##### Determiner le fichier excel 

input("\nAppuyer sur Entree pour choisir le fichier excel comportant la base de donnee avec les resp de groupes\n")

Tk().withdraw() 
contenu1 = askopenfilename()

xl = pd.ExcelFile(contenu1)

v_sheet = 0

if len(xl.sheet_names) > 1 :
	print ("\n---Les differents feuilles presentes---")
	for x in range (len(xl.sheet_names)) :
		print (str(x)+" "+xl.sheet_names[x])

	v_sheet = int(input("Entrer le numero de la feuille excel\n"))

	
#######################################################################################################
##### Determiner les lignes prises en compte

reponse = input("\nTaper 1 si le script doit prendre en compte l'ensemble des lignes des destinataires.\nTaper 2 pour indiquer quelles lignes doivent etre prises en compte\n")

if reponse=='2':
	vligned = int(input("Numero de la ligne de debut de la liste des destinataires "))
	vlignef = int(input("Numero de la ligne de fin de la liste des destinataires "))
else:
	vligned = 1
	vlignef = 1
	

#######################################################################################################
##### Definir les colonnes


vdept = parser.get('colonnes', 'dept')
vgrp = parser.get('colonnes', 'grp')

print ("Lettre colonne departement "+vdept)
print ("Lettre colonne groupe "+vgrp)


print ("\n----Verifier que les informations ci dessus sont correctes----")
time.sleep(1)

while 1:
	
	reponse = input("\nTaper 1 pour modifier manuellement les colonnes prises en compte, sinon taper 2\n")

	if reponse=='1':
	
		vdept = input("Indiquer la lettre de la colonne comportant le departement\n")
		vdept = fct2 (vdept)
		vgrp = input("Indiquer la lettre de la colonne comportant le groupe\n")
		vgrp = fct2 (vgrp)
		break
		
	elif reponse == '2' :
		
		vdept = fct2 (vdept)
		vgrp = fct2 (vgrp)
		break

	else :

		print ("Choix incorrect !")
		#break

#######################################################################################################
##### Definir les responsables de groupe


vresp = input("Indiquer la lettre de la colonne cochee indiquant les responsables de groupe\n")
vresp = fct2 (vresp)



#########################################################################
# Convertion fichier excel

vnbrligne = ((vlignef+1)-vligned)

read_file = pd.read_excel(contenu1,sheet_name=v_sheet,skiprows = vligned-1, header=None)         

read_file.to_csv ("Test.csv",  
                  index = None, 
                  header = True)################uft8 a faire

#print (read_file)##########

df = pd.DataFrame(pd.read_csv("Test.csv")) 
vrespa = df.columns.values[vresp] 
vdepta = df.columns.values[vdept] 

if vlignef != 1 :
	df = df[:vnbrligne]
	pd.DataFrame(df.dropna(subset=[vdepta], inplace=True))
else :
	pd.DataFrame(df.dropna(subset=[vdepta], inplace=True))

#print (df)############

pd.DataFrame(df.dropna(subset=[vrespa], inplace=True)) ######suppression des non-resp
df[vresp] = df[vrespa].str.strip()#### suppression espace colonne resp
df = df[(df[vrespa].str.match('X'))|(df[vrespa].str.match('x'))]#### suppression resp sans X

list = []

for x in range (1,len(df.index)+1):
 list.append(x)

df.index = list

os.remove("Test.csv")


#######################################################################################################
##### Verification

i = 0

t = fct3 (vgrp, 'groupe', 0)

4
while t[0] == 1 :
	i = i+1
	z = t[1]
	t = fct3 (vgrp, 'groupe', z)


if i > 0 :
	print (str(i)+" erreurs")
	time.sleep (1)
	sys.exit(0)


print (df) ################################


#######################################################################################################
##### ouverture commande

input("\nAppuyer sur Entree pour choisir le fichier excel de commande\n")

Tk().withdraw() 
contenu2 = askopenfilename()

xl2 = pd.ExcelFile(contenu2)

v_sheet2 = 0

if len(xl2.sheet_names) > 1 :
	print ("\n---Les differents feuilles presentes---")
	for x in range (len(xl2.sheet_names)) :
		print (str(x)+" "+xl2.sheet_names[x])

	reponse = int(input("Entrer le numero de la feuille excel\n"))
	v_sheet2 = xl2.sheet_names[reponse]

#######################################################################################################
##### Creation dico grp-dept

dico = dict()

for x in range (len(df.index)):
	dico.update( {df.iat[x,vgrp] : df.iat[x,vdept]} )

print (dico)	
print (len(df.index))

#######################################################################################################
##### modif case et creation fichiers

ns = input("entrer le numero de semaine\n")
wb2 = load_workbook(filename = contenu2)

if len(xl2.sheet_names) > 1 :
	ws2 = wb2[v_sheet2]
	
else :
	ws2 = wb2.active
	
os.chdir(dossier_usr)
	
for key in dico : 
	d = dico.get(key)
	
	if d.isdigit() and int(d) < 10 :
		d = "0"+d
		ws2.cell(row = 2, column = 5).value = int(d)
		
		for values, keys in dico.items():
			if "0"+keys == d :
				# ~ print (d)
				e = values
				
	else :
		ws2.cell(row = 2, column = 5).value = d
	
		for values, keys in dico.items():
			if keys == d :
				# ~ print (d)
				e = values
				
	wb2.save ("cmd_"+d+"_"+e+"_S"+ns+".xlsx")
	wb2.close()

